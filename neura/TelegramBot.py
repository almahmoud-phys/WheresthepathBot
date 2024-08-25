import asyncio
import datetime
import os
import random
import shutil

# openpyxl
import openpyxl
from async_tkinter_loop import async_handler
from openpyxl.styles import Font
from pytgcalls import PyTgCalls
from telethon import utils

import neura.Constants as Constants
import neura.utils as ut


class TelegramBot:
    def __init__(self, client, group, users):
        """
        Initializes the TelegramBot instance.

        :param client: The Telethon client instance for interacting with Telegram.
        :param group: The group object that the bot is interacting with.
        :param users: A dictionary containing user information, with user IDs as keys.
        """
        self.loop = asyncio.get_event_loop()
        self.client = client
        self.group = group
        self.users = users

    # all the methods that are going to be called from the main thread should be decorated with @async_handler
    # so they don't block the main thread

    @async_handler
    async def send_message(self, user_ids: list, message: str, callback):
        """
        Sends a message to a list of users asynchronously. It introduces a random delay between
        messages to avoid spamming.

        :param user_ids: A list of user IDs or usernames to send the message to.
        :param message: The message to be sent.
        :param callback: A callback function that is called when all messages are sent.
        """
        print("Sending message")
        for user_id in user_ids:
            try:
                # Check if '@' is in user_id, then split
                if "@" in user_id:
                    recipient_id = user_id.split("@")[1]
                else:
                    recipient_id = user_id

                print(f"Sending message to {recipient_id}")

                # Send the message to the recipient
                await self.client.send_message(recipient_id, message)

                # Introduce a random delay between sending messages
                await asyncio.sleep(random.uniform(30, 60))
            except Exception as e:
                print(f"Failed to send message to {user_id}: {str(e)}")

        # Update the UI and invoke the callback function on finish
        callback("Messages sent")
        ut.show_info("Messages sent")

    async def group_info(self):
        pass

    @async_handler
    async def take_presence(self, step, callback):
        """
        Tracks the presence of users in a Telegram voice chat and saves the attendance
        to an Excel file. The file is updated or created using a template if it doesn't exist.

        :param step: An integer representing the step الخطوة number (used as a column in the Excel file).
        :param callback: A callback function that is called when the presence tracking is complete.
        """
        print("Taking presence in bot")
        chat_id = utils.get_peer_id(self.group)

        call_py = PyTgCalls(self.client)
        await call_py.start()
        global running
        running = True
        list_to_save = []
        while running:
            participants = await call_py.get_participants(chat_id)
            callback(f"{len(participants)}" + " participants in the voice chat")
            for participant in participants:
                id = participant.user_id
                list_to_save.append(id)
            if len(participants) == 0:
                running = False
            await asyncio.sleep(1)
        ut.show_info("Presence taking done")
        print("Presence taking done")
        # remove duplicates
        list_to_save = list(dict.fromkeys(list_to_save))

        # use default directory to save the file
        # if the file doesn't exist, it will be created from template
        # TODO: also group name should be the file name
        file_path = Constants.DEFAULT_DIRICTORY + "presence.xlsx"
        if not os.path.exists(file_path):
            shutil.copy("./assets/presence_template.xlsx", file_path)

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # group title
        sheet["B1"] = self.group.title

        print(f"Step {step}")
        step = int(step)
        # column_number is the step number + 3
        column_number = step + 6

        # get today date
        sheet.cell(2, column_number).value = datetime.datetime.now().strftime(
            "%Y-%m-%d"
        )
        sheet.cell(3, column_number).value = f"Step {step}"

        row_number = 3  # The specific row where we start adding presence of users

        for id in list_to_save:
            user = self.users[id]
            row_number += 1
            sheet.cell(row_number, 1).value = user["id"]
            sheet.cell(
                row_number, 2
            ).value = f"{user["first_name"]} {user["last_name"]}"
            sheet.cell(row_number, 3).value = user["username"]
            sheet.cell(row_number, column_number).value = 1

        # save the workbook
        workbook.save(file_path)

        # Update the UI and invoke the callback function on finish
        callback(f"Presence saved to {file_path}")

    @async_handler
    async def export_users(self, callback):
        """
        Exports users information from the group to an Excel file. The file is created using
        a template if it doesn't exist, and it includes a hyperlink to each user's Telegram profile.

        :param callback: A callback function that is called when the export is complete.
        """
        # TODO : file name should be group name
        # use default directory to save the file
        file_path = Constants.DEFAULT_DIRICTORY + "users_info.xlsx"

        # open xlsx file or copy default template from assets it if it doesn't exist
        if not os.path.exists(file_path):
            shutil.copy("./assets/users_info_template.xlsx", file_path)

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # group title
        sheet["B1"] = self.group.title

        # get today date
        sheet["B2"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # group participants count
        sheet["B3"] = self.group.participants_count

        row_number = 5  # The specific row where we start adding users data
        # Add user data
        for id in self.users:
            user = self.users[id]

            # the first cell in the row is the user id and hyperlink to the user profile in telegram to make it easy to contact the user
            hyperlink_cell = sheet.cell(row=row_number, column=1)
            hyperlink_cell.value = user["id"]
            hyperlink_cell.hyperlink = f'https://web.telegram.org/a/#{user["id"]}'
            # Apply blue color and underline font style
            hyperlink_cell.font = Font(color="0000FF", underline="single")

            user_data = [
                f"{user["first_name"]} {user["last_name"]}",
                user["username"],
                0,  # Placeholder for user interaction in the group
                user["is_admin"],
                user["safe_to_send"],
                "",  # Placeholder for note about the user , if the user change name or username  , quit the group .. or any other note
            ]
            for col, value in enumerate(user_data, start=2):
                sheet.cell(row=row_number, column=col, value=value)
            row_number += 1

        workbook.save(file_path)
        callback(f"Users exported to {file_path}")
        ut.show_info(f"Users exported to {file_path}")
